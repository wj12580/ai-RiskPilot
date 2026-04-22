"""
导出服务：生成 Excel 报告
"""

import io
import json
import base64
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime


def _apply_table_style(ws, start_row, start_col, end_row, end_col, header_fill='2563EB'):
    """应用表格样式：交替行颜色"""
    light_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    header_fill_obj = PatternFill(start_color=header_fill, end_color=header_fill, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    cell_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if row == start_row:
                cell.fill = header_fill_obj
                cell.font = header_font
            else:
                cell.fill = light_fill if (row - start_row) % 2 == 1 else white_fill
                cell.font = cell_font


def export_analysis_report(task_dict: dict) -> bytes:
    """
    导出分析任务报告为 Excel（增强版：包含AI建议、图表等）
    """
    result = task_dict.get('result', {})
    summary = result.get('summary', {})
    metrics = result.get('metrics', {})
    bins = result.get('bins', [])
    feature = result.get('feature_analysis', {})
    
    # 智能解析 suggestion（可能是 JSON 字符串、列表、字典等多种格式）
    suggestion = task_dict.get('suggestion', [])
    
    # 如果是字符串，尝试解析为JSON
    if isinstance(suggestion, str):
        try:
            suggestion = json.loads(suggestion)
        except:
            suggestion = []
    
    # 如果是字典（Agent模式），尝试提取其中的文本内容
    if isinstance(suggestion, dict):
        # Agent模式：{'type': 'agent', 'report': {...}}
        if suggestion.get('type') == 'agent':
            report = suggestion.get('report', {})
            final_report = report.get('final_report', '')
            if final_report:
                suggestion = [{
                    'level': 'info',
                    'category': 'AI专家分析',
                    'title': 'Agent分析报告',
                    'content': final_report,
                    'action': '请参考上述分析内容进行策略优化'
                }]
            else:
                suggestion = []
        # multi_model模式：{'type': 'multi_model', 'html': True}
        elif suggestion.get('type') == 'multi_model':
            suggestion = [{
                'level': 'info',
                'category': '多模型分析',
                'title': '多模型综合分析已完成',
                'content': '详见HTML报告中的相关性热力图、聚类树状图、互补性矩阵、串行策略模拟结果',
                'action': '查看图表区域获取详细分析'
            }]
        # multi_expert模式：{'type': 'multi_expert', 'expert_reports': {...}, 'final_report': str}
        elif suggestion.get('type') == 'multi_expert':
            expert_reports = suggestion.get('expert_reports', {})
            final_report = suggestion.get('final_report', '')
            suggestions_list = suggestion.get('suggestions', [])
            
            if suggestions_list and isinstance(suggestions_list, list):
                suggestion = suggestions_list
            elif final_report:
                suggestion = [{
                    'level': 'info',
                    'category': '多专家分析',
                    'title': '综合分析报告',
                    'content': final_report[:5000],  # 限制长度
                    'action': '详见完整报告'
                }]
            else:
                suggestion = []
        else:
            suggestion = []
    
    # 确保是列表
    if suggestion is None:
        suggestion = []
    
    # 根据 analysis_type 判断分析模式
    analysis_type = task_dict.get('analysis_type', '')
    analysis_tags = task_dict.get('analysis_tags', '')
    feature_cols = task_dict.get('feature_cols', '')
    
    # 判断是否是 Agent 模式
    mode = task_dict.get('mode', 'traditional')
    if not mode:
        # 从 analysis_type 和 feature_cols 推断
        if 'agent' in str(analysis_type).lower():
            mode = 'agent'
        elif feature_cols and ',' in feature_cols:
            mode = 'multi_model'
        elif suggestion and isinstance(suggestion, dict) and suggestion.get('type') == 'multi_expert':
            mode = 'multi_expert'
        else:
            mode = 'traditional'

    wb = Workbook()

    # ── Sheet 1: 概览 ─────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = '分析概览'
    
    # 设置列宽
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 25

    # 标题
    ws1['A1'] = 'RiskPilot 策略分析报告'
    ws1['A1'].font = Font(size=18, bold=True, color='FFFFFF')
    ws1['A1'].fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.merge_cells('A1:D1')
    ws1.row_dimensions[1].height = 40

    # 报告信息
    ws1['A3'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws1['A3'].font = Font(size=10, color='6B7280')

    # 基础信息
    info_start = 5
    ws1[f'A{info_start}'] = '📋 基本信息'
    ws1[f'A{info_start}'].font = Font(size=14, bold=True, color='2563EB')

    info_rows = [
        ['文件名称', task_dict.get('file_name', '-')],
        ['分析模式', '多模型分析' if mode == 'multi_model' else ('多专家分析' if mode == 'multi_expert' else ('Agent分析' if mode == 'agent' else '传统分析'))],
        ['分析类型', analysis_tags.replace(',', ' | ') if analysis_tags else '-'],
        ['目标列', task_dict.get('target_col', '-')],
        ['分数列', task_dict.get('score_col', '-')],
        ['分箱数', str(task_dict.get('n_bins', 10))],
        ['用户备注', task_dict.get('user_note', '-')],
    ]
    
    for i, (k, v) in enumerate(info_rows, start=info_start + 1):
        ws1[f'A{i}'] = k
        ws1[f'B{i}'] = v
        ws1[f'A{i}'].font = Font(bold=True)
        ws1[f'A{i}'].fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')

    # 汇总统计
    summary_start = info_start + len(info_rows) + 2
    ws1[f'A{summary_start}'] = '📊 汇总统计'
    ws1[f'A{summary_start}'].font = Font(size=14, bold=True, color='2563EB')

    summary_rows = [
        ['样本总数', f"{summary.get('total_count', 0):,}"],
        ['坏样本数', f"{summary.get('bad_count', 0):,}"],
        ['逾期率', f"{summary.get('bad_rate', 0):.2%}"],
    ]
    
    for i, (k, v) in enumerate(summary_rows, start=summary_start + 1):
        ws1[f'A{i}'] = k
        ws1[f'B{i}'] = v
        ws1[f'A{i}'].font = Font(bold=True)
        ws1[f'C{i}'] = v
        ws1[f'C{i}'].font = Font(bold=True, size=14, color='DC2626')

    # 模型指标
    metrics_start = summary_start + len(summary_rows) + 2
    ws1[f'A{metrics_start}'] = '🎯 模型指标'
    ws1[f'A{metrics_start}'].font = Font(size=14, bold=True, color='2563EB')

    metric_rows = [
        ['KS 值', f"{metrics.get('ks', 0):.4f}", _get_indicator_text(metrics.get('ks', 0), 0.35, 0.25)],
        ['AUC', f"{metrics.get('auc', 0):.4f}", _get_indicator_text(metrics.get('auc', 0), 0.75, 0.65)],
        ['PSI', f"{metrics.get('psi', 0):.4f}", _get_psi_text(metrics.get('psi', 0))],
    ]
    
    for i, (k, v, indicator) in enumerate(metric_rows, start=metrics_start + 1):
        ws1[f'A{i}'] = k
        ws1[f'B{i}'] = v
        ws1[f'C{i}'] = indicator
        ws1[f'A{i}'].font = Font(bold=True)
        
        # 颜色指示
        if '✓' in indicator:
            ws1[f'C{i}'].font = Font(color='059669', bold=True)
        elif '⚠' in indicator:
            ws1[f'C{i}'].font = Font(color='D97706', bold=True)
        else:
            ws1[f'C{i}'].font = Font(color='DC2626', bold=True)

    # ── Sheet 2: 分箱详情 ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('分箱详情')
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 15

    if bins:
        # 标题行
        headers = ['分箱序号', '分数下限', '分数上限', '样本数', '坏样本数', '逾期率(%)', '逾期率']
        for col, header in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=header)
        
        for idx, bin_data in enumerate(bins):
            row = idx + 2
            ws2.cell(row=row, column=1, value=bin_data.get('bin_no', idx + 1))
            ws2.cell(row=row, column=2, value=round(bin_data.get('score_min', 0), 4))
            ws2.cell(row=row, column=3, value=round(bin_data.get('score_max', 0), 4))
            ws2.cell(row=row, column=4, value=bin_data.get('count', 0))
            ws2.cell(row=row, column=5, value=bin_data.get('bad_count', 0))
            ws2.cell(row=row, column=6, value=round(bin_data.get('bad_rate', 0) * 100, 2))
            ws2.cell(row=row, column=7, value=bin_data.get('bad_rate', 0))
        
        _apply_table_style(ws2, 1, 1, len(bins) + 1, 7)

    # ── Sheet 3: AI策略建议 ────────────────────────────────────────────────────
    ws3 = wb.create_sheet('AI策略建议')
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 50
    ws3.column_dimensions['D'].width = 50

    # 标题
    ws3['A1'] = '💡 AI 策略建议'
    ws3['A1'].font = Font(size=16, bold=True, color='2563EB')
    ws3.merge_cells('A1:D1')

    # 增强版建议解析：支持多种格式
    parsed_suggestions = []
    
    # 统一处理：如果是字符串，尝试解析
    if isinstance(suggestion, str):
        try:
            suggestion = json.loads(suggestion)
        except:
            suggestion = []
    
    if suggestion and isinstance(suggestion, list) and len(suggestion) > 0:
        # 情况1：标准建议格式 [{level, category, title, content, action}]
        first_item = suggestion[0]
        if isinstance(first_item, dict) and 'title' in first_item:
            parsed_suggestions = suggestion
        # 情况2：Agent报告格式 [{content, action, category, level}]
        elif isinstance(first_item, dict) and 'content' in first_item:
            for s in suggestion:
                # 检查是否有长文本的final_report
                final_report = s.get('final_report', '')
                content = s.get('content', '')
                
                # 如果有final_report，用它作为content
                if final_report:
                    content = final_report
                
                if content:  # 只处理有内容的建议
                    parsed_suggestions.append({
                        'level': s.get('level', 'info'),
                        'category': s.get('category', 'AI分析'),
                        'title': s.get('title', s.get('category', 'AI分析')) or 'AI分析',
                        'content': content,
                        'action': s.get('action', ''),
                    })
    
    # 情况3：直接从task_dict中提取final_report（Agent模式）
    if not parsed_suggestions:
        agent_report = task_dict.get('agent_report', {})
        if isinstance(agent_report, dict):
            final_report = agent_report.get('final_report', '')
            if final_report:
                parsed_suggestions = [{
                    'level': 'info',
                    'category': 'AI专家分析',
                    'title': 'AI分析报告摘要',
                    'content': final_report,
                    'action': '请参考上述分析内容进行策略优化'
                }]
    
    if parsed_suggestions:
        # 表头
        headers = ['优先级', '类别', '建议标题', '建议内容', '操作建议']
        for col, header in enumerate(headers, 1):
            ws3.cell(row=3, column=col, value=header)
        
        # 级别颜色映射
        level_colors = {
            'high': 'DC2626',   # 红色
            'medium': 'D97706', # 橙色
            'low': '059669',    # 绿色
            'info': '2563EB',   # 蓝色
            'warning': 'D97706',
            'success': '059669',
        }
        level_map = {
            'high': '🔴 高', 
            'medium': '🟡 中', 
            'low': '🟢 低', 
            'info': '🔵 信息',
            'warning': '⚠️ 警告',
            'success': '✅ 良好'
        }
        
        for idx, s in enumerate(parsed_suggestions):
            row = idx + 4
            level = s.get('level', 'info')
            level_text = level_map.get(level, '🔵 信息')
            
            ws3.cell(row=row, column=1, value=level_text)
            ws3.cell(row=row, column=1).font = Font(color=level_colors.get(level, '2563EB'), bold=True)
            
            ws3.cell(row=row, column=2, value=s.get('category', '-'))
            ws3.cell(row=row, column=3, value=s.get('title', '-'))
            ws3.cell(row=row, column=4, value=s.get('content', '-'))
            ws3.cell(row=row, column=5, value=s.get('action', '-'))
        
        _apply_table_style(ws3, 3, 1, len(parsed_suggestions) + 3, 5)
    else:
        ws3['A3'] = '暂无AI策略建议'

    # ── Sheet 4: 特征分析（如果有）────────────────────────────────────────────
    if feature:
        ws4 = wb.create_sheet('特征分析')
        ws4.column_dimensions['A'].width = 20
        ws4.column_dimensions['B'].width = 25

        feature_rows = [
            ['特征列', feature.get('score_col', '-')],
            ['目标列', feature.get('target_col', '-')],
            ['均值', f"{feature.get('mean', 0):.4f}"],
            ['标准差', f"{feature.get('std', 0):.4f}"],
            ['最小值', f"{feature.get('min', 0):.4f}"],
            ['最大值', f"{feature.get('max', 0):.4f}"],
            ['中位数', f"{feature.get('median', 0):.4f}"],
            ['25分位', f"{feature.get('p25', 0):.4f}"],
            ['75分位', f"{feature.get('p75', 0):.4f}"],
        ]
        
        for i, (k, v) in enumerate(feature_rows, start=1):
            ws4.cell(row=i, column=1, value=k)
            ws4.cell(row=i, column=2, value=v)
            ws4.cell(row=i, column=1).font = Font(bold=True)
            ws4.cell(row=i, column=1).fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')

    # ── Sheet 5: 多模型分析结果（如果是多模型模式）────────────────────────────
    if mode == 'multi_model' and result.get('performance'):
        ws5 = wb.create_sheet('多模型对比')
        ws5.column_dimensions['A'].width = 20
        ws5.column_dimensions['B'].width = 15
        ws5.column_dimensions['C'].width = 15
        ws5.column_dimensions['D'].width = 15
        ws5.column_dimensions['E'].width = 15

        perf = result.get('performance', [])
        headers = ['模型名称', 'KS', 'AUC', '逾期率', '样本量']
        for col, header in enumerate(headers, 1):
            ws5.cell(row=1, column=col, value=header)
        
        for idx, p in enumerate(perf):
            row = idx + 2
            ws5.cell(row=row, column=1, value=p.get('model', '-'))
            ws5.cell(row=row, column=2, value=f"{p.get('ks', 0):.4f}")
            ws5.cell(row=row, column=3, value=f"{p.get('auc', 0):.4f}")
            ws5.cell(row=row, column=4, value=f"{p.get('bad_rate', 0):.2%}")
            ws5.cell(row=row, column=5, value=p.get('count', 0))
        
        _apply_table_style(ws5, 1, 1, len(perf) + 1, 5)

    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _get_indicator_text(value, good_thresh, warn_thresh):
    """获取指标评估文本"""
    if value >= good_thresh:
        return f'✓ 良好 (≥{good_thresh})'
    elif value >= warn_thresh:
        return f'⚠ 一般 ({warn_thresh}~{good_thresh})'
    else:
        return f'✗ 较差 (<{warn_thresh})'


def _get_psi_text(psi):
    """获取PSI评估文本"""
    if psi < 0.1:
        return '✓ 稳定 (<0.1)'
    elif psi < 0.25:
        return '⚠ 轻微波动 (0.1~0.25)'
    else:
        return '✗ 显著变化 (≥0.25)'


def export_records_excel(records: list) -> bytes:
    """
    导出策略调整记录为 Excel
    """
    if not records:
        wb = Workbook()
        ws = wb.active
        ws.title = '策略调整记录'
        ws['A1'] = '暂无记录'
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # 构造 DataFrame
    rows = []
    for r in records:
        rows.append({
            'ID':           r.get('id'),
            '策略名称':     r.get('strategy_name'),
            '调整日期':     r.get('adjusted_at'),
            '策略类型':     r.get('strategy_type'),
            '调整内容':     r.get('content'),
            '调整原因':     ', '.join(r.get('reason_tags', [])),
            '预期目标':     r.get('expected_goal'),
            '复盘状态':     r.get('review_status'),
            '备注':         r.get('notes'),
        })

    df = pd.DataFrame(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = '策略调整记录'

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
